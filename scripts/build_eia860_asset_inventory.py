#!/usr/bin/env python3
"""Build a first EIA-860 plant/generator inventory for the EOP012 rebuild.

This script intentionally stops before weather matching. The goal is to make the
asset universe visible and auditable before any ECWT calculation work begins.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import sys
import zipfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import openpyxl

from eop012_config import EIA860_ZIP, PROJECT_ROOT

DEFAULT_ZIP = EIA860_ZIP
DEFAULT_PROJECT_ROOT = PROJECT_ROOT


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value).strip()
    text = text.replace("\n", " ")
    text = re.sub(r"[^0-9A-Za-z]+", "_", text).strip("_").lower()
    return text


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_note_or_footer_row(values: list[str]) -> bool:
    nonblank = [value for value in values if value]
    if not nonblank:
        return True
    first = nonblank[0].upper()
    return first.startswith("NOTE:") or first.startswith("SOURCE:")


def find_member(members: Iterable[str], pattern: str) -> str:
    regex = re.compile(pattern, re.IGNORECASE)
    matches = [name for name in members if regex.search(Path(name).name)]
    if not matches:
        raise ValueError(f"No ZIP member matched {pattern!r}")
    if len(matches) > 1:
        raise ValueError(f"Multiple ZIP members matched {pattern!r}: {matches}")
    return matches[0]


def load_workbook_from_zip(zip_path: Path, member: str):
    with zipfile.ZipFile(zip_path) as zf:
        data = zf.read(member)
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def find_header_row(ws, required_headers: set[str]) -> tuple[int, list[str], list[str]]:
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        raw_headers = [clean_cell(v) for v in row]
        headers = [normalize_header(v) for v in raw_headers]
        if required_headers.issubset(set(headers)):
            return row_idx, raw_headers, headers
    raise ValueError(f"Could not find header row in sheet {ws.title!r}; required {required_headers}")


def read_table(wb, sheet_name: str, required_headers: set[str], source_table: str) -> tuple[list[str], list[dict[str, str]]]:
    ws = wb[sheet_name]
    header_row, raw_headers, headers = find_header_row(ws, required_headers)

    # Preserve duplicate source columns by suffixing normalized names.
    seen: Counter[str] = Counter()
    unique_headers: list[str] = []
    for header in headers:
        if not header:
            header = "unnamed"
        seen[header] += 1
        unique_headers.append(header if seen[header] == 1 else f"{header}_{seen[header]}")

    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        values = [clean_cell(v) for v in row[: len(unique_headers)]]
        if is_note_or_footer_row(values):
            continue
        item = {key: values[idx] if idx < len(values) else "" for idx, key in enumerate(unique_headers)}
        item["_source_table"] = source_table
        rows.append(item)
    return unique_headers + ["_source_table"], rows


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def parse_float(value: str) -> float | None:
    try:
        if value == "":
            return None
        return float(value)
    except ValueError:
        return None


def counter_table(counter: Counter[str], limit: int | None = None) -> str:
    rows = counter.most_common(limit)
    if not rows:
        return "_None._\n"
    lines = ["| Value | Count |", "| --- | ---: |"]
    for value, count in rows:
        lines.append(f"| `{value or '[blank]'}` | {count} |")
    return "\n".join(lines) + "\n"


def infer_release_label(zip_path: Path) -> str:
    stem = zip_path.stem
    label = re.sub(r"[^0-9A-Za-z]+", "_", stem).strip("_").lower()
    return label


def build_inventory(zip_path: Path, project_root: Path) -> dict:
    if not zip_path.exists():
        raise FileNotFoundError(zip_path)

    release_label = infer_release_label(zip_path)
    processed_dir = project_root / "data" / "processed" / release_label
    docs_dir = project_root / "docs"
    processed_dir.mkdir(parents=True, exist_ok=True)
    docs_dir.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(zip_path) as zf:
        bad_member = zf.testzip()
        if bad_member:
            raise ValueError(f"ZIP integrity check failed at member {bad_member}")
        members = zf.namelist()
        member_info = {
            info.filename: {
                "file_name": Path(info.filename).name,
                "size_bytes": info.file_size,
                "compressed_size_bytes": info.compress_size,
                "modified_at_zip": "%04d-%02d-%02dT%02d:%02d:%02d"
                % (*info.date_time,),
            }
            for info in zf.infolist()
        }

    plant_member = find_member(members, r"^2_+Plant_.*\.xlsx$")
    generator_member = find_member(members, r"^3_1_Generator_.*\.xlsx$")
    utility_member = find_member(members, r"^1_+Utility_.*\.xlsx$")
    layout_member = find_member(members, r"^Layout.*\.xlsx$")

    plant_wb = load_workbook_from_zip(zip_path, plant_member)
    plant_fields, plant_rows = read_table(
        plant_wb,
        "Plant",
        {"plant_code", "plant_name", "latitude", "longitude", "state"},
        "Plant",
    )

    utility_wb = load_workbook_from_zip(zip_path, utility_member)
    utility_fields, utility_rows = read_table(
        utility_wb,
        "Utility",
        {"utility_id", "utility_name"},
        "Utility",
    )

    generator_wb = load_workbook_from_zip(zip_path, generator_member)
    generator_outputs: dict[str, tuple[list[str], list[dict[str, str]]]] = {}
    for sheet in generator_wb.sheetnames:
        source_table = f"Generator / {sheet}"
        fields, rows = read_table(
            generator_wb,
            sheet,
            {"plant_code", "plant_name", "generator_id", "status"},
            source_table,
        )
        generator_outputs[sheet] = (fields, rows)

    layout_wb = load_workbook_from_zip(zip_path, layout_member)
    field_fields, field_rows = read_table(
        layout_wb,
        "Field Directory",
        {"field_name", "description"},
        "Field Directory",
    )

    write_csv(processed_dir / "plant.csv", plant_rows, plant_fields)
    write_csv(processed_dir / "utility.csv", utility_rows, utility_fields)
    write_csv(processed_dir / "field_dictionary.csv", field_rows, field_fields)

    generator_summary: dict[str, dict] = {}
    all_generator_rows: list[dict[str, str]] = []
    all_generator_fields: list[str] = []
    for sheet, (fields, rows) in generator_outputs.items():
        file_stem = normalize_header(sheet)
        out_name = f"generator_{file_stem}.csv"
        write_csv(processed_dir / out_name, rows, fields)
        all_generator_rows.extend(rows)
        for field in fields:
            if field not in all_generator_fields:
                all_generator_fields.append(field)
        generator_summary[sheet] = {
            "rows": len(rows),
            "distinct_plants": len({r.get("plant_code", "") for r in rows if r.get("plant_code", "")}),
            "status_counts": Counter(r.get("status", "") for r in rows),
            "technology_counts": Counter(r.get("technology", "") for r in rows),
            "prime_mover_counts": Counter(r.get("prime_mover", "") for r in rows),
            "missing_plant_code": sum(1 for r in rows if not r.get("plant_code")),
            "missing_generator_id": sum(1 for r in rows if not r.get("generator_id")),
        }
    write_csv(processed_dir / "generator_all.csv", all_generator_rows, all_generator_fields)

    plant_codes = {r.get("plant_code", "") for r in plant_rows if r.get("plant_code", "")}
    operable_rows = generator_outputs.get("Operable", ([], []))[1]
    operable_plant_codes = {r.get("plant_code", "") for r in operable_rows if r.get("plant_code", "")}
    all_generator_plant_codes = {r.get("plant_code", "") for r in all_generator_rows if r.get("plant_code", "")}

    coord_valid = 0
    coord_missing = 0
    coord_invalid = 0
    suspicious_coords: list[dict[str, str]] = []
    for row in plant_rows:
        lat = parse_float(row.get("latitude", ""))
        lon = parse_float(row.get("longitude", ""))
        if lat is None or lon is None:
            coord_missing += 1
            continue
        if -90 <= lat <= 90 and -180 <= lon <= 180:
            coord_valid += 1
        else:
            coord_invalid += 1
            suspicious_coords.append(row)

    plant_state_counts = Counter(r.get("state", "") for r in plant_rows)
    plant_sector_counts = Counter(r.get("sector_name", "") for r in plant_rows)
    operable_status_counts = generator_summary.get("Operable", {}).get("status_counts", Counter())

    source_manifest = [
        {
            "source_path": str(zip_path),
            "file_name": zip_path.name,
            "source_type": "eia860_zip",
            "size_bytes": zip_path.stat().st_size,
            "sha256": sha256_file(zip_path),
            "processed_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        }
    ]
    for member, info in member_info.items():
        if member in {plant_member, generator_member, utility_member, layout_member}:
            source_manifest.append(
                {
                    "source_path": f"{zip_path}!/{member}",
                    "file_name": info["file_name"],
                    "source_type": "zip_member",
                    "size_bytes": info["size_bytes"],
                    "sha256": "",
                    "processed_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                }
            )
    write_csv(
        processed_dir / "source_manifest.csv",
        source_manifest,
        ["source_path", "file_name", "source_type", "size_bytes", "sha256", "processed_at_utc"],
    )

    summary = {
        "zip_path": str(zip_path),
        "release_label": release_label,
        "processed_dir": str(processed_dir),
        "zip_sha256": source_manifest[0]["sha256"],
        "members_used": {
            "plant": plant_member,
            "generator": generator_member,
            "utility": utility_member,
            "layout": layout_member,
        },
        "plant_rows": len(plant_rows),
        "utility_rows": len(utility_rows),
        "field_dictionary_rows": len(field_rows),
        "coordinate_counts": {
            "valid_numeric": coord_valid,
            "missing": coord_missing,
            "invalid_range": coord_invalid,
        },
        "distinct_plant_codes": len(plant_codes),
        "plants_with_operable_generators": len(operable_plant_codes),
        "plant_rows_without_operable_generator": len(plant_codes - operable_plant_codes),
        "operable_generator_rows": len(operable_rows),
        "generator_rows_all_sheets": len(all_generator_rows),
        "generator_plant_codes_not_in_plant_table": sorted(all_generator_plant_codes - plant_codes),
        "operable_plant_codes_not_in_plant_table": sorted(operable_plant_codes - plant_codes),
        "plant_state_counts": dict(plant_state_counts),
        "plant_sector_counts": dict(plant_sector_counts),
        "operable_status_counts": dict(operable_status_counts),
        "generator_summary": {
            sheet: {
                "rows": data["rows"],
                "distinct_plants": data["distinct_plants"],
                "status_counts": dict(data["status_counts"]),
                "technology_counts_top_20": dict(data["technology_counts"].most_common(20)),
                "prime_mover_counts_top_20": dict(data["prime_mover_counts"].most_common(20)),
                "missing_plant_code": data["missing_plant_code"],
                "missing_generator_id": data["missing_generator_id"],
            }
            for sheet, data in generator_summary.items()
        },
    }

    (processed_dir / "inventory_summary.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

    report = render_markdown_report(summary, plant_state_counts, plant_sector_counts, generator_summary)
    (docs_dir / "eia860_asset_inventory.md").write_text(report, encoding="utf-8")
    (docs_dir / "data_inventory.md").write_text(report, encoding="utf-8")

    return summary


def render_markdown_report(
    summary: dict,
    plant_state_counts: Counter[str],
    plant_sector_counts: Counter[str],
    generator_summary: dict[str, dict],
) -> str:
    lines: list[str] = []
    lines.append("# EIA-860 Asset Inventory")
    lines.append("")
    lines.append(f"Generated UTC: {datetime.now(timezone.utc).isoformat(timespec='seconds')}")
    lines.append("")
    lines.append("## Source")
    lines.append("")
    lines.append(f"- ZIP: `{summary['zip_path']}`")
    lines.append(f"- Release label: `{summary['release_label']}`")
    lines.append(f"- SHA-256: `{summary['zip_sha256']}`")
    lines.append("- Baseline use: 2024 final annual data should be treated as the current authoritative baseline until final 2025 EIA-860 data is published.")
    lines.append("")
    lines.append("## ZIP Members Used")
    lines.append("")
    for key, value in summary["members_used"].items():
        lines.append(f"- {key}: `{value}`")
    lines.append("")
    lines.append("## Core Counts")
    lines.append("")
    lines.append("| Metric | Count |")
    lines.append("| --- | ---: |")
    lines.append(f"| Plant rows | {summary['plant_rows']} |")
    lines.append(f"| Distinct plant codes | {summary['distinct_plant_codes']} |")
    lines.append(f"| Utility rows | {summary['utility_rows']} |")
    lines.append(f"| Field dictionary rows | {summary['field_dictionary_rows']} |")
    lines.append(f"| Generator rows, all sheets | {summary['generator_rows_all_sheets']} |")
    lines.append(f"| Operable generator rows | {summary['operable_generator_rows']} |")
    lines.append(f"| Plants with at least one operable generator | {summary['plants_with_operable_generators']} |")
    lines.append(f"| Plant rows without operable generator | {summary['plant_rows_without_operable_generator']} |")
    lines.append("")
    lines.append("## Coordinate Quality")
    lines.append("")
    lines.append("| Coordinate check | Count |")
    lines.append("| --- | ---: |")
    coord = summary["coordinate_counts"]
    lines.append(f"| Valid numeric lat/lon | {coord['valid_numeric']} |")
    lines.append(f"| Missing/non-numeric lat/lon | {coord['missing']} |")
    lines.append(f"| Outside valid lat/lon range | {coord['invalid_range']} |")
    lines.append("")
    lines.append("## Generator Sheets")
    lines.append("")
    lines.append("| Sheet | Rows | Distinct plants | Missing plant code | Missing generator ID |")
    lines.append("| --- | ---: | ---: | ---: | ---: |")
    for sheet, data in generator_summary.items():
        lines.append(
            f"| {sheet} | {data['rows']} | {data['distinct_plants']} | {data['missing_plant_code']} | {data['missing_generator_id']} |"
        )
    lines.append("")
    lines.append("## Operable Generator Status Counts")
    lines.append("")
    operable = generator_summary.get("Operable")
    if operable:
        lines.append(counter_table(operable["status_counts"]))
    else:
        lines.append("_No Operable sheet found._\n")
    lines.append("## Plant Sector Counts")
    lines.append("")
    lines.append(counter_table(plant_sector_counts))
    lines.append("## Plant State Counts")
    lines.append("")
    lines.append(counter_table(plant_state_counts))
    lines.append("## Top Operable Technologies")
    lines.append("")
    if operable:
        lines.append(counter_table(operable["technology_counts"], limit=30))
    else:
        lines.append("_No Operable sheet found._\n")
    lines.append("## Top Operable Prime Movers")
    lines.append("")
    if operable:
        lines.append(counter_table(operable["prime_mover_counts"], limit=30))
    else:
        lines.append("_No Operable sheet found._\n")
    lines.append("## Reconciliation Checks")
    lines.append("")
    gen_missing = summary["generator_plant_codes_not_in_plant_table"]
    op_missing = summary["operable_plant_codes_not_in_plant_table"]
    lines.append(f"- Generator plant codes not found in plant table: {len(gen_missing)}")
    if gen_missing:
        lines.append(f"  - First 25: `{', '.join(gen_missing[:25])}`")
    lines.append(f"- Operable plant codes not found in plant table: {len(op_missing)}")
    if op_missing:
        lines.append(f"  - First 25: `{', '.join(op_missing[:25])}`")
    lines.append("")
    lines.append("## ECWT Implications")
    lines.append("")
    lines.append("- Use `Plant Code` as the primary plant key for station matching.")
    lines.append("- Use plant latitude/longitude for initial NOAA station candidate generation.")
    lines.append("- Use the Operable generator sheet as the first current equipment universe; refine EOP-012 applicability later.")
    lines.append("- Keep retired/canceled and proposed units out of the main current ECWT output, but retain them as separate auditable layers.")
    lines.append("- Before ECWT calculation, every included plant must have a coordinate, selected representative weather station path, expected-hour count, valid-hour count, and missing/excess-data flag.")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--zip", type=Path, default=DEFAULT_ZIP, help="Path to the EIA-860 ZIP file.")
    parser.add_argument(
        "--project-root",
        type=Path,
        default=DEFAULT_PROJECT_ROOT,
        help="Path to the rebuild project root.",
    )
    args = parser.parse_args()

    summary = build_inventory(args.zip, args.project_root)
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
